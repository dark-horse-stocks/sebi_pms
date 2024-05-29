import pickle
import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook


def get_portfolio_manager_general_information(portfolio_manager_value, portfolio_manager_name, year, month):
    month_dict = {
        1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
        7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"
    }

    # Find the key for the given month name
    month_key = None
    for key, value in month_dict.items():
        if value == month:
            month_key = key
            break

    url = f'https://www.sebi.gov.in/sebiweb/other/OtherAction.do?doPmr=yes&pmrId={portfolio_manager_value}&year={int(year)}&month={int(month_key)}'
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')

    try:
        portfolio_manager = soup.find(
            'th', text='Name of the Portfolio Manager').find_next('td').text.strip().upper()
    except AttributeError:
        portfolio_manager = portfolio_manager_name

    try:
        clients_count = soup.find(
            'th', text='No. of clients as on last day of the month').find_next('td').text.strip()
    except AttributeError:
        clients_count = 0

    try:
        aum = soup.find('th', text='Total Assets under Management (AUM) as on last day of the month (Amount in INR crores)').find_next(
            'td').text.strip()
    except AttributeError:
        aum = 0

    number_of_clients_field = f'{month} - Clients'
    aum_field = f'{month} - AUM'

    portfolio_manager_data = {
        'Portfolio Manager': portfolio_manager,
        "Year": year,
        number_of_clients_field: float(clients_count),
        aum_field: float(aum)
    }

    return portfolio_manager_data


def check_portfolio_manager_past_months(df, options, year, month, portfolio_managers):

    month_dict = {
        1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
        7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"
    }
    # Find the key for the given month name
    month_key = None
    for key, value in month_dict.items():
        if value == month:
            month_key = key
            break

    if month_key >= 4:
        end = 4
    elif month_key == 3:
        end = 3
    elif month_key == 2:
        end = 2
    else:
        return df

    data_list = []
    if portfolio_managers:
        # Create a reverse dictionary
        reverse_dict = {v: k for k, v in options.items()}
        for portfolio_manager in portfolio_managers:
            # Find the key for the desired value using the reverse dictionary
            key = reverse_dict.get(portfolio_manager)
            for i in range(1, end):
                data = get_portfolio_manager_general_information(
                    key, portfolio_manager, str(year), month_dict[month_key - i])
                data_list.append(data)

    if data_list:
        for item in data_list:
            keys = list(item.keys())
            values = list(item.values())
            year = f"{values[1]}"
            month = f"{keys[2].split('-')[0].strip()}"
            if values[-1] != 0:
                df.loc[values[0], (year, month, 'Clients')] = values[2]
                df.loc[values[0], (year, month, 'AUM')] = values[3]

    return df


def add_new_portfolio_manager_data(year, month):

    year = str(year)
    month = str(month)
    # To load the dictionary object from the file
    with open('options.pkl', 'rb') as file:
        options = pickle.load(file)

    portfolio_managers_data = []

    for index, (key, value) in enumerate(options.items(), start=1):
        progress_percentage = (index / len(options)) * 100
        print(
            f'Processing {index}/{len(options)} ({progress_percentage:.2f}%)')

        portfolio_manager_data = get_portfolio_manager_general_information(
            key, value, year, month)

        portfolio_managers_data.append(portfolio_manager_data)

    # Convert the list of dictionaries to a pandas DataFrame
    df1 = pd.DataFrame(portfolio_managers_data)

    # Extract relevant columns from the DataFrame
    Clients = [float(value) for value in df1[f'{month} - Clients']]
    AUM = [float(value) for value in df1[f'{month} - AUM']]

    df = pd.read_excel('SEBI-Portfolio-Manager-General-Information.xlsx',
                       header=[0, 1, 2], index_col=[0], sheet_name='Master')

    if (year, month, 'AUM') in df.columns:
        before = df[(year, month, 'AUM')]

        df[(year, month, 'Clients')] = Clients
        df[(year, month, 'AUM')] = AUM

        after = df[(year, month, 'AUM')]

        # Find the differences between 'before' and 'after'
        differences = after - before

        # Get the portfolio managers where changes occurred
        changed_managers = differences[differences != 0]

        portfolio_managers = []
        for index, value in changed_managers.items():
            portfolio_managers.append(index)

        if portfolio_managers and len(portfolio_managers) < 444:
            df = check_portfolio_manager_past_months(
                df, options, year, month, list(set(portfolio_managers)))
    else:
        df.insert(2, (year, month, 'Clients'), Clients)
        df.insert(3, (year, month, 'AUM'), AUM)

    return df


def get_Particulars_data(portfolio_manager_value, portfolio_manager_name, year, month):

    month_dict = {
        1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
        7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"
    }
    # Find the key for the given month name
    month_key = None
    for key, value in month_dict.items():
        if value == month:
            month_key = key
            break

    url = f'https://www.sebi.gov.in/sebiweb/other/OtherAction.do?doPmr=yes&pmrId={portfolio_manager_value}&year={int(year)}&month={int(month_key)}'
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')

    try:
        # Find the table element
        table = soup.find('th', text='Domestic Clients').find_parent(
            'table', class_='table table-striped table-bordered table-hover background statistics-table')

        # Extract data from the first and second rows of the tbody
        tbody_rows = table.find("tbody").find_all("tr")

        # Extract data for the first row
        first_row_data = [portfolio_manager_name, 'No. Clients']
        for cell in tbody_rows[0].find_all(["th", "td"])[1:]:
            first_row_data.append(cell.text.strip())

        # Extract data for the second row
        second_row_data = [portfolio_manager_name, 'AUM']
        for cell in tbody_rows[1].find_all(["th", "td"])[1:]:
            second_row_data.append(cell.text.strip())

    except AttributeError:
        first_row_data = [portfolio_manager_name, 'No. Clients'] + [0] * 7
        second_row_data = [portfolio_manager_name, 'AUM'] + [0] * 7

    data = [first_row_data, second_row_data]
    # Transforming the list to the desired format
    output = []

    Domestic_Clients_PF = f'{month} - Domestic Clients PF/EPFO'
    Domestic_Clients_Corporates = f'{month} - Domestic Clients Corporates'
    Domestic_Clients_Non_Corporates = f'{month} - Domestic Clients Non-Corporates'
    Foreign_Clients_Non_Residents = f'{month} - Foreign Clients Non-Residents'
    Foreign_Clients_FPI = f'{month} - Foreign Clients FPI'
    Foreign_Clients_Others = f'{month} - Foreign Clients Others'
    Total = f'{month} - Total'

    for item in data:
        output_dict = {
            "Portfolio Manager": item[0],
            "Particulars": item[1],
            "Year": year,
            Domestic_Clients_PF: float(item[2]),
            Domestic_Clients_Corporates: float(item[3]),
            Domestic_Clients_Non_Corporates: float(item[4]),
            Foreign_Clients_Non_Residents: float(item[5]),
            Foreign_Clients_FPI: float(item[6]),
            Foreign_Clients_Others: float(item[7]),
            Total: float(item[8])
        }
        output.append(output_dict)

    return output[0], output[1]


def check_particulars_past_months(df, options, year, month, portfolio_managers):

    month_dict = {
        1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
        7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"
    }
    # Find the key for the given month name
    month_key = None
    for key, value in month_dict.items():
        if value == month:
            month_key = key
            break

    if month_key >= 4:
        end = 4
    elif month_key == 3:
        end = 3
    elif month_key == 2:
        end = 2
    else:
        return df

    data_list = []
    if portfolio_managers:
        # Create a reverse dictionary
        reverse_dict = {v: k for k, v in options.items()}
        for portfolio_manager in portfolio_managers:
            # Find the key for the desired value using the reverse dictionary
            key = reverse_dict.get(portfolio_manager)
            for i in range(1, end):
                data1, data2 = get_Particulars_data(
                    key, portfolio_manager, str(year), month_dict[month_key - i])
                data_list.append(data1)
                data_list.append(data2)

    if data_list:
        for item in data_list:
            keys = list(item.keys())
            values = list(item.values())
            year = f"{values[2]}"
            month = f"{keys[3].split('-')[0].strip()}"
            if values[-1] != 0:
                if values[1] == 'No. Clients':
                    df.loc[(df.index == values[0]) & (df[(' ', ' ', 'Particulars', ' ')] ==
                                                      'No. Clients'), (year, month, 'Domestic Clients', 'PF/EPFO')] = values[3]
                    df.loc[(df.index == values[0]) & (df[(' ', ' ', 'Particulars', ' ')] ==
                                                      'No. Clients'), (year, month, 'Domestic Clients', 'Corporates')] = values[4]
                    df.loc[(df.index == values[0]) & (df[(' ', ' ', 'Particulars', ' ')] ==
                                                      'No. Clients'), (year, month, 'Domestic Clients', 'Non-Corporates')] = values[5]
                    df.loc[(df.index == values[0]) & (df[(' ', ' ', 'Particulars', ' ')] ==
                                                      'No. Clients'), (year, month, 'Foreign Clients', 'Non-Residents')] = values[6]
                    df.loc[(df.index == values[0]) & (df[(' ', ' ', 'Particulars', ' ')] ==
                                                      'No. Clients'), (year, month, 'Foreign Clients', 'FPI')] = values[7]
                    df.loc[(df.index == values[0]) & (df[(' ', ' ', 'Particulars', ' ')] ==
                                                      'No. Clients'), (year, month, 'Foreign Clients', 'Others')] = values[8]
                    df.loc[(df.index == values[0]) & (df[(' ', ' ', 'Particulars', ' ')]
                                                      == 'No. Clients'), (year, month, ' ', 'Total')] = values[9]
                else:
                    df.loc[(df.index == values[0]) & (df[(' ', ' ', 'Particulars', ' ')]
                                                      == 'AUM'), (year, month, 'Domestic Clients', 'PF/EPFO')] = values[3]
                    df.loc[(df.index == values[0]) & (df[(' ', ' ', 'Particulars', ' ')] ==
                                                      'AUM'), (year, month, 'Domestic Clients', 'Corporates')] = values[4]
                    df.loc[(df.index == values[0]) & (df[(' ', ' ', 'Particulars', ' ')] ==
                                                      'AUM'), (year, month, 'Domestic Clients', 'Non-Corporates')] = values[5]
                    df.loc[(df.index == values[0]) & (df[(' ', ' ', 'Particulars', ' ')] ==
                                                      'AUM'), (year, month, 'Foreign Clients', 'Non-Residents')] = values[6]
                    df.loc[(df.index == values[0]) & (df[(' ', ' ', 'Particulars', ' ')]
                                                      == 'AUM'), (year, month, 'Foreign Clients', 'FPI')] = values[7]
                    df.loc[(df.index == values[0]) & (df[(' ', ' ', 'Particulars', ' ')]
                                                      == 'AUM'), (year, month, 'Foreign Clients', 'Others')] = values[8]
                    df.loc[(df.index == values[0]) & (
                        df[(' ', ' ', 'Particulars', ' ')] == 'AUM'), (year, month, ' ', 'Total')] = values[9]

    return df


def add_new_particulars_data(year, month):

    year = str(year)
    month = str(month)
    # To load the dictionary object from the file
    with open('options.pkl', 'rb') as file:
        options = pickle.load(file)

    portfolio_managers_data = []

    for index, (key, value) in enumerate(options.items(), start=1):
        progress_percentage = (index / len(options)) * 100
        print(
            f'Processing {index}/{len(options)} ({progress_percentage:.2f}%)')

        result_0, result_1 = get_Particulars_data(key, value, year, month)

        portfolio_managers_data.append(result_0)
        portfolio_managers_data.append(result_1)

    # Convert the list of dictionaries to a pandas DataFrame
    df1 = pd.DataFrame(portfolio_managers_data)

    # Extract relevant columns from the DataFrame
    EPFO = [float(value)
            for value in df1[f'{month} - Domestic Clients PF/EPFO']]
    Corporates = [float(value)
                  for value in df1[f'{month} - Domestic Clients Corporates']]
    Non_Corporates = [
        float(value) for value in df1[f'{month} - Domestic Clients Non-Corporates']]
    Non_Residents = [
        float(value) for value in df1[f'{month} - Foreign Clients Non-Residents']]
    FPI = [float(value) for value in df1[f'{month} - Foreign Clients FPI']]
    Others = [float(value)
              for value in df1[f'{month} - Foreign Clients Others']]
    Total = [float(value) for value in df1[f'{month} - Total']]

    df = pd.read_excel('SEBI-Portfolio-Manager-Particulars.xlsx',
                       header=[0, 1, 2, 3], index_col=[0], sheet_name='Master')

    if (year, month, ' ', 'Total') in df.columns:
        before = df[(year, month, ' ', 'Total')]

        df[(year, month, 'Domestic Clients', 'PF/EPFO')] = EPFO
        df[(year, month, 'Domestic Clients', 'Corporates')] = Corporates
        df[(year, month, 'Domestic Clients', 'Non-Corporates')] = Non_Corporates
        df[(year, month, 'Foreign Clients', 'Non-Residents')] = Non_Residents
        df[(year, month, 'Foreign Clients', 'FPI')] = FPI
        df[(year, month, 'Foreign Clients', 'Others')] = Others
        df[(year, month, ' ', 'Total')] = Total

        after = df[(year, month, ' ', 'Total')]

        # Find the differences between 'before' and 'after'
        differences = after - before

        # Get the portfolio managers where changes occurred
        changed_managers = differences[differences != 0]

        portfolio_managers = []
        # Print the portfolio managers where changes occurred
        for index, value in changed_managers.items():
            portfolio_managers.append(index)

        if portfolio_managers and len(portfolio_managers) < 444:
            df = check_particulars_past_months(
                df, options, year, month, list(set(portfolio_managers)))
    else:
        df.insert(1, (year, month, 'Domestic Clients', 'PF/EPFO'), EPFO)
        df.insert(2, (year, month, 'Domestic Clients', 'Corporates'), Corporates)
        df.insert(3, (year, month, 'Domestic Clients',
                  'Non-Corporates'), Non_Corporates)
        df.insert(4, (year, month, 'Foreign Clients',
                  'Non-Residents'), Non_Residents)
        df.insert(5, (year, month, 'Foreign Clients', 'FPI'), FPI)
        df.insert(6, (year, month, 'Foreign Clients', 'Others'), Others)
        df.insert(7, (year, month, ' ', 'Total'), Total)

    return df


def get_Investment_data(portfolio_manager_value, portfolio_manager_name, year, month):
    month_dict = {
        1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
        7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec"
    }

    # Find the key for the given month name
    month_key = None
    for key, value in month_dict.items():
        if value == month:
            month_key = key
            break

    url = f'https://www.sebi.gov.in/sebiweb/other/OtherAction.do?doPmr=yes&pmrId={portfolio_manager_value}&year={int(year)}&month={int(month_key)}'
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')

    Equity_Listed = f'{year}-{month}-Listed'
    Equity_Unlisted = f'{year}-{month}-Unlisted'
    Mutual_Funds = f'{year}-{month}-Mutual Funds'
    Others = f'{year}-{month}-Others'
    Total = f'{year}-{month}-Total'

    try:
        # Find the table element
        table = soup.find('th', text='Investment Approach').find_parent(
            'table', class_='table table-striped table-bordered table-hover background statistics-table')

        # Extract data from the first and second rows of the tbody
        tbody_rows = table.find("tbody").find_all("tr")[:-1]

        data = []
        for row in tbody_rows:
            row_data = [portfolio_manager_name]
            for cell in row.find_all(["th", "td"]):
                row_data.append(cell.text.strip())

            data.append({
                "Portfolio Manager": row_data[0],
                "Investment Approach": row_data[1],
                Equity_Listed: float(row_data[2]),
                Equity_Unlisted: float(row_data[3]),
                Mutual_Funds: float(row_data[-3]),
                Others: float(row_data[-2]),
                Total: float(row_data[-1])
            })

    except AttributeError:
        data = [{
                "Portfolio Manager": portfolio_manager_name,
                "Investment Approach": "No Investment Approach",
                Equity_Listed: 0,
                Equity_Unlisted: 0,
                Mutual_Funds: 0,
                Others: 0,
                Total: 0
                }]

    return data


def add_new_investment_data(year, month):
    # To load the dictionary object from the file
    with open('options.pkl', 'rb') as file:
        options = pickle.load(file)

    year = str(year)
    portfolio_managers_data = []

    for index, (key, value) in enumerate(options.items(), start=1):
        progress_percentage = (index / len(options)) * 100
        print(
            f'Processing {index}/{len(options)} ({progress_percentage:.2f}%)')

        portfolio_manager_data = get_Investment_data(
            key, value, year, str(month))

        for item in portfolio_manager_data:
            portfolio_managers_data.append(item)

    # Formatting the Investment Approach column
    def fix_string(string):
        words = string.split()
        words = [word.capitalize() for word in words]
        return ' '.join(words)

    # Convert the list of dictionaries to a pandas DataFrame
    df1 = pd.DataFrame(portfolio_managers_data)
    df1['Investment Approach'] = df1['Investment Approach'].apply(fix_string)

    # Set the multi-index
    df1.set_index(['Portfolio Manager', 'Investment Approach'], inplace=True)

    df0 = pd.read_excel('SEBI-Portfolio-Manager-Investment-Approach.xlsx',
                        header=[0, 1, 2, 3, 4], sheet_name='Master', index_col=[0, 1])

    # Convert MultiIndex to single level
    df0_new_columns = []
    for col in df0.columns:
        year = col[0]
        month = col[1]
        category = col[4] if col[4].strip() else col[3]

        df0_new_columns.append(f"{year}-{month}-{category}")

    df0.columns = df0_new_columns

    df = df0.join(df1, how='outer')

    df_new_columns = []
    for col in df.columns:
        year_month, category = col.rsplit('-', 1)
        year, month = year_month.split('-')
        # Adjust for Mutual Funds, Others, and Total
        if category in ['Listed', 'Unlisted']:
            equity = 'Equity'
            sub_category = category
        else:
            equity = category
            sub_category = ' '

        df_new_columns.append(
            (year, month, '(AUM) as on last day of the month (Amount in INR crores)', equity, sub_category))

    # Create the MultiIndex
    multi_index = pd.MultiIndex.from_tuples(df_new_columns)

    df.columns = multi_index

    return df


# Functions
# Function to process the Portfolio Manager General Information Excel file
def process_general_information_excel(file_path):
    # Load your DataFrame
    df = pd.read_excel(file_path, header=[0, 1, 2], index_col=[
                       0], sheet_name='Master')

    # Create new column names by joining the levels of the MultiIndex
    new_columns = [' '.join(filter(None, col)).strip() for col in df.columns]

    # Assign the new column names to the DataFrame
    df.columns = new_columns

    # Reset the index to convert it to a column
    df.reset_index(inplace=True)

    # Define month mapping
    month_map = {
        'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04',
        'May': '05', 'Jun': '06', 'Jul': '07', 'Aug': '08',
        'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'
    }

    def convert_date_format(col_name):
        parts = col_name.split()
        if len(parts) == 2 and parts[0].isdigit() and parts[1] in month_map:
            year = parts[0]
            month = month_map[parts[1]]
            return f"{year}/{month}"
        return col_name

    # Filter and process columns for Clients
    clients_df = df[[col for col in df.columns if 'AUM' not in col]]
    clients_df.columns = [convert_date_format(
        col.replace(' Clients', '')) for col in clients_df.columns]

    # Filter and process columns for AUM
    aum_df = df[[col for col in df.columns if 'Clients' not in col]]
    aum_df.columns = [convert_date_format(
        col.replace(' AUM', '')) for col in clients_df.columns]

    # Load the workbook
    workbook = load_workbook(file_path)

    # Function to remove existing sheet
    def remove_sheet_if_exists(sheet_name):
        if sheet_name in workbook.sheetnames:
            std = workbook[sheet_name]
            workbook.remove(std)
            workbook.save(file_path)  # Save after removing the sheet

    # Remove existing sheets if they exist
    remove_sheet_if_exists('Clients')
    remove_sheet_if_exists('AUM')

    # Save the filtered DataFrames to new sheets in the same Excel file
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        clients_df.to_excel(writer, sheet_name='Clients', index=False)
        aum_df.to_excel(writer, sheet_name='AUM', index=False)


# Function to process the Portfolio Manager Particulars Excel file
def split_particulars_columns(file_path, df, client_type):
    # Define month mapping
    month_map = {
        'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04',
        'May': '05', 'Jun': '06', 'Jul': '07', 'Aug': '08',
        'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'
    }

    def convert_date_format(col_name):
        parts = col_name.strip().split()
        if len(parts) == 2 and parts[0].isdigit() and parts[1] in month_map:
            year = parts[0]
            month = month_map[parts[1]]
            return f"{year}/{month}"
        return col_name

    # Domestic Clients
        # Process PF/EPFO Column
    domestic_clients_PF_EPFO_df = df[[
        col for col in df.columns if 'PF/EPFO' in col or col in ['Portfolio Manager']]]
    domestic_clients_PF_EPFO_df.columns = [convert_date_format(col.replace(
        ' Domestic Clients PF/EPFO', '')) for col in domestic_clients_PF_EPFO_df.columns]
    # Process Corporates Column
    domestic_clients_Corporates_df = df[[col for col in df.columns if col.endswith(
        'Clients Corporates') or col in ['Portfolio Manager']]]
    domestic_clients_Corporates_df.columns = [convert_date_format(col.replace(
        ' Domestic Clients Corporates', '')) for col in domestic_clients_Corporates_df.columns]
    # Process Non-Corporates Column
    domestic_clients_Non_Corporates_df = df[[
        col for col in df.columns if 'Non-Corporates' in col or col in ['Portfolio Manager']]]
    domestic_clients_Non_Corporates_df.columns = [convert_date_format(col.replace(
        ' Domestic Clients Non-Corporates', '')) for col in domestic_clients_Non_Corporates_df.columns]
    # Foreign Clients
    # Process Non-Residents Column
    foreign_clients_Non_Residents_df = df[[
        col for col in df.columns if 'Non-Residents' in col or col in ['Portfolio Manager']]]
    foreign_clients_Non_Residents_df.columns = [convert_date_format(col.replace(
        ' Foreign Clients Non-Residents', '')) for col in foreign_clients_Non_Residents_df.columns]
    # Process FPI Column
    foreign_clients_FPI_df = df[[
        col for col in df.columns if 'FPI' in col or col in ['Portfolio Manager']]]
    foreign_clients_FPI_df.columns = [convert_date_format(col.replace(
        ' Foreign Clients FPI', '')) for col in foreign_clients_FPI_df.columns]
    # Process Others Column
    foreign_clients_Others_df = df[[
        col for col in df.columns if 'Others' in col or col in ['Portfolio Manager']]]
    foreign_clients_Others_df.columns = [convert_date_format(col.replace(
        ' Foreign Clients Others', '')) for col in foreign_clients_Others_df.columns]
    # Total
    total_df = df[[col for col in df.columns if 'Total' in col or col in [
        'Portfolio Manager']]]
    total_df.columns = [convert_date_format(
        col.replace(' Total', '')) for col in total_df.columns]

    # Load the workbook
    workbook = load_workbook(file_path)

    # Function to remove existing sheet
    def remove_sheet_if_exists(sheet_name):
        if sheet_name in workbook.sheetnames:
            std = workbook[sheet_name]
            workbook.remove(std)
            workbook.save(file_path)  # Save after removing the sheet

    # Remove existing sheets if they exist
    remove_sheet_if_exists(f'{client_type} Domestic Clients PF_EPFO')
    remove_sheet_if_exists(f'{client_type} Domestic Clients Corporates')
    remove_sheet_if_exists(f'{client_type} Domestic Clients Non_Corporates')
    remove_sheet_if_exists(f'{client_type} Foreign Clients Non_Residents')
    remove_sheet_if_exists(f'{client_type} Foreign Clients FPI')
    remove_sheet_if_exists(f'{client_type} Foreign Clients Others')
    remove_sheet_if_exists(f'{client_type} Total')

    # Save the filtered DataFrames to new sheets in the same Excel file
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        domestic_clients_PF_EPFO_df.to_excel(
            writer, sheet_name=f'{client_type} Domestic Clients PF_EPFO', index=False)
        domestic_clients_Corporates_df.to_excel(
            writer, sheet_name=f'{client_type} Domestic Clients Corporates', index=False)
        domestic_clients_Non_Corporates_df.to_excel(
            writer, sheet_name=f'{client_type} Domestic Clients Non_Corporates', index=False)
        foreign_clients_Non_Residents_df.to_excel(
            writer, sheet_name=f'{client_type} Foreign Clients Non_Residents', index=False)
        foreign_clients_FPI_df.to_excel(
            writer, sheet_name=f'{client_type} Foreign Clients FPI', index=False)
        foreign_clients_Others_df.to_excel(
            writer, sheet_name=f'{client_type} Foreign Clients Others', index=False)
        total_df.to_excel(
            writer, sheet_name=f'{client_type} Total', index=False)


def process_particulars_excel(file_path):
    # Load the DataFrame
    df = pd.read_excel(file_path, header=[0, 1, 2, 3], index_col=[
                       0], sheet_name='Master')

    # Create new column names by joining the levels of the MultiIndex
    new_columns = [' '.join(filter(None, col)).strip() for col in df.columns]
    df.columns = new_columns

    # Reset the index to convert it to a column
    df.reset_index(inplace=True)

    # Filter DataFrame based on 'Particulars' column
    clients_df = df[df['Particulars'] == 'No. Clients'].copy()
    aum_df = df[df['Particulars'] == 'AUM'].copy()

    # Filter and process columns for Clients
    split_particulars_columns(file_path, clients_df, 'Clients')

    # Filter and process columns for AUM
    split_particulars_columns(file_path, aum_df, 'AUM')


# Function to process the Portfolio Manager Investment Approach Excel file
def split_investment_columns(file_path, df, client_type):
    # Define month mapping
    month_map = {
        'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04',
        'May': '05', 'Jun': '06', 'Jul': '07', 'Aug': '08',
        'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12'
    }

    def convert_date_format(col_name):
        parts = col_name.strip().split()
        if len(parts) == 2 and parts[0].isdigit() and parts[1] in month_map:
            year = parts[0]
            month = month_map[parts[1]]
            return f"{year}/{month}"
        return col_name

    # Equity
        # Process Listed Column
    equity_listed_df = df[[col for col in df.columns if 'Listed' in col or col in [
        'Portfolio Manager', 'Investment Approach']]]
    equity_listed_df.columns = [convert_date_format(col.replace(
        ' Equity Listed', '')) for col in equity_listed_df.columns]
    # Process Unlisted Column
    equity_unlisted_df = df[[col for col in df.columns if 'Unlisted' in col or col in [
        'Portfolio Manager', 'Investment Approach']]]
    equity_unlisted_df.columns = [convert_date_format(col.replace(
        ' Equity Unlisted', '')) for col in equity_unlisted_df.columns]
    # Process Mutual Funds Column
    mutual_funds_df = df[[col for col in df.columns if 'Mutual Funds' in col or col in [
        'Portfolio Manager', 'Investment Approach']]]
    mutual_funds_df.columns = [convert_date_format(col.replace(
        ' Mutual Funds', '')) for col in mutual_funds_df.columns]
    # Process Others Column
    others_df = df[[col for col in df.columns if 'Others' in col or col in [
        'Portfolio Manager', 'Investment Approach']]]
    others_df.columns = [convert_date_format(
        col.replace(' Others', '')) for col in others_df.columns]
    # Process Total Column
    total_df = df[[col for col in df.columns if 'Total' in col or col in [
        'Portfolio Manager', 'Investment Approach']]]
    total_df.columns = [convert_date_format(
        col.replace(' Total', '')) for col in total_df.columns]

    # Load the workbook
    workbook = load_workbook(file_path)

    # Function to remove existing sheet
    def remove_sheet_if_exists(sheet_name):
        if sheet_name in workbook.sheetnames:
            std = workbook[sheet_name]
            workbook.remove(std)
            workbook.save(file_path)  # Save after removing the sheet

    # Remove existing sheets if they exist
    remove_sheet_if_exists(f'{client_type} Equity Listed')
    remove_sheet_if_exists(f'{client_type} Equity Unlisted')
    remove_sheet_if_exists(f'{client_type} Mutual Funds')
    remove_sheet_if_exists(f'{client_type} Others')
    remove_sheet_if_exists(f'{client_type} Total')

    # Save the filtered DataFrames to new sheets in the same Excel file
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        equity_listed_df.set_index(
            ['Portfolio Manager', 'Investment Approach'], inplace=True)
        equity_listed_df.to_excel(
            writer, sheet_name=f'{client_type} Equity Listed')
        equity_unlisted_df.set_index(
            ['Portfolio Manager', 'Investment Approach'], inplace=True)
        equity_unlisted_df.to_excel(
            writer, sheet_name=f'{client_type} Equity Unlisted')
        mutual_funds_df.set_index(
            ['Portfolio Manager', 'Investment Approach'], inplace=True)
        mutual_funds_df.to_excel(
            writer, sheet_name=f'{client_type} Mutual Funds')
        others_df.set_index(
            ['Portfolio Manager', 'Investment Approach'], inplace=True)
        others_df.to_excel(writer, sheet_name=f'{client_type} Others')
        total_df.set_index(
            ['Portfolio Manager', 'Investment Approach'], inplace=True)
        total_df.to_excel(writer, sheet_name=f'{client_type} Total')


def process_investment_excel(file_path):
    # Load the DataFrame
    df = pd.read_excel(file_path, header=[0, 1, 2, 3, 4], index_col=[
                       0, 1], sheet_name='Master')

    # Create new column names by joining the levels of the MultiIndex
    new_columns = [' '.join(filter(None, col)).strip() for col in df.columns]
    new_columns = [col.replace(
        ' (AUM) as on last day of the month (Amount in INR crores)', '') for col in new_columns]
    df.columns = new_columns

    # Reset the index to convert it to a column
    df.reset_index(inplace=True)

    # Filter and process columns for AUM
    split_investment_columns(file_path, df, 'AUM')

# Main Functions


# For SEBI-Portfolio-Manager-General-Information.xlsx
# Add Sep Data for the general innformation
def portfolio_manager_data(year, month):
    file_path = 'SEBI-Portfolio-Manager-General-Information.xlsx'
    df = add_new_portfolio_manager_data(year, month)
    df.to_excel(file_path)
    print('Please wait while the data is being processed...')
    process_general_information_excel(file_path)


# For SEBI-Portfolio-Manager-Particulars.xlsx
# Add Sep Data for the Particulars
def particulars_data(year, month):
    file_path = 'SEBI-Portfolio-Manager-Particulars.xlsx'
    df = add_new_particulars_data(year, month)
    df.to_excel(file_path)
    print('Please wait while the data is being processed...')
    process_particulars_excel(file_path)


# For SEBI-Portfolio-Manager-Investment-Approach.xlsx
# Add Sep Data for the Investment
def investment_data(year, month):
    file_path = 'SEBI-Portfolio-Manager-Investment-Approach.xlsx'
    df = add_new_investment_data(year, month)
    df.to_excel(file_path)
    print('Please wait while the data is being processed...')
    process_investment_excel(file_path)
